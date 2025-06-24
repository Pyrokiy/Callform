import sys
import os
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import (
    QApplication, QWidget, QLineEdit, QPushButton,
    QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QFileDialog, QMessageBox, QDateEdit, QListWidget, QListWidgetItem, QComboBox, QLayout
)
from PyQt5.QtCore import QDate, Qt
from openpyxl import Workbook, load_workbook


class MultiSelectComboBox(QComboBox):
    def __init__(self, items, parent=None):
        super().__init__(parent)
        self.setEditable(True)
        self.lineEdit().setReadOnly(True)
        self.setInsertPolicy(QComboBox.NoInsert)
        self.setPlaceholderText("申請区分を選択")

        self.list_widget = QListWidget()
        self.setModel(self.list_widget.model())
        self.setView(self.list_widget)

        for item_text in items:
            item = QListWidgetItem(item_text)
            item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Unchecked)
            self.list_widget.addItem(item)

        self.list_widget.itemChanged.connect(self.update_text)

    def setPlaceholderText(self, text):
        self.lineEdit().setPlaceholderText(text)

    def update_text(self):
        selected = []
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            if item.checkState() == Qt.Checked:
                selected.append(item.text())
        self.lineEdit().setText(", ".join(selected))

    def selected_items(self):
        return [self.list_widget.item(i).text()
                for i in range(self.list_widget.count())
                if self.list_widget.item(i).checkState() == Qt.Checked]

    def clear_selection(self):
        for i in range(self.list_widget.count()):
            self.list_widget.item(i).setCheckState(Qt.Unchecked)
        self.lineEdit().clear()


class LoanFormApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setFont(QFont("Meiryo", 14))
        self.setWindowTitle("発送依頼アプリ")
        self.setGeometry(200, 200, 1000, 600)
        self.excel_path = None
        self.sheet_name = "Sheet1"
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        font = QFont("Meiryo", 14)

        self.date_input = QDateEdit()
        self.date_input.setFont(font)
        self.date_input.setCalendarPopup(True)
        self.date_input.setDate(QDate.currentDate())

        self.staff_input = QLineEdit()
        self.staff_input.setFont(font)
        self.staff_input.setPlaceholderText("担当者名")

        self.request_type = MultiSelectComboBox([
            "通常免除申請書", "猶予申告書", "住所・氏名 変更届",
            "払込票（月賦）", "任意免除申請書", "変額申請書(少額・増額・一括口振)",
            "口振依頼書", "状況申告状況", "その他"])
        self.request_type.setFont(font)

        reset_request_type_btn = QPushButton("リセット")
        reset_request_type_btn.setFont(font)
        reset_request_type_btn.clicked.connect(self.request_type.clear_selection)

        request_type_layout = QHBoxLayout()
        request_type_layout.addWidget(self.request_type)
        request_type_layout.addWidget(reset_request_type_btn)

        self.request_detail = QLineEdit()
        self.request_detail.setFont(font)
        self.request_detail.setPlaceholderText("申請詳細")

        self.borrower_input = QLineEdit()
        self.borrower_input.setFont(font)
        self.borrower_input.setPlaceholderText("借受人氏名（ひらがな→Enterでカタカナ）")
        self.borrower_input.returnPressed.connect(self.convert_to_katakana)

        self.code_input = QLineEdit()
        self.code_input.setFont(font)
        self.code_input.setPlaceholderText("貸付コード")

        self.note_input = QLineEdit()
        self.note_input.setFont(font)
        self.note_input.setPlaceholderText("特記事項")
        self.note_input.setMinimumWidth(300)

        input_layout = QHBoxLayout()
        left_layout = QVBoxLayout()
        right_layout = QVBoxLayout()

        for w in [self.date_input, self.staff_input, request_type_layout, self.request_detail]:
            if isinstance(w, QLayout):
                left_layout.addLayout(w)
            else:
                left_layout.addWidget(w)

        for w in [self.borrower_input, self.code_input, self.note_input]:
            right_layout.addWidget(w)

        input_layout.addLayout(left_layout)
        input_layout.addLayout(right_layout)

        self.load_button = QPushButton("参照ファイルを開く")
        self.load_button.setFont(font)
        self.load_button.clicked.connect(self.load_existing_excel)

        self.add_button = QPushButton("追加（入力内容を追加します。）")
        self.add_button.setFont(font)
        self.add_button.clicked.connect(self.add_entry)

        self.delete_button = QPushButton("削除（間違えた行の数字を選択してから削除）")
        self.delete_button.setFont(font)
        self.delete_button.clicked.connect(self.delete_entry)

        self.save_button = QPushButton("保存（入力・変更した内容を保存）")
        self.save_button.setFont(font)
        self.save_button.clicked.connect(self.save_data)

        button_layout = QHBoxLayout()
        button_layout.addWidget(self.add_button)
        button_layout.addWidget(self.delete_button)
        button_layout.addWidget(self.save_button)

        self.table = QTableWidget()
        self.table.setFont(font)
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels([
            "日付", "担当者名", "申請区分", "申請詳細",
            "借受人氏名", "貸付コード", "特記事項"])
        self.table.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.SelectedClicked)

        layout.addWidget(self.load_button)
        layout.addLayout(input_layout)
        layout.addLayout(button_layout)
        layout.addWidget(self.table)

        self.setLayout(layout)

    def convert_to_katakana(self):
        text = self.borrower_input.text()
        katakana = ''.join(
            chr(ord(char) + 96) if 'ぁ' <= char <= 'ん' else char for char in text
        )
        self.borrower_input.setText(katakana)

    def add_entry(self):
        row_position = self.table.rowCount()
        self.table.insertRow(row_position)
        values = [
            self.date_input.date().toString("yyyy/MM/dd"),
            self.staff_input.text(),
            ", ".join(self.request_type.selected_items()),
            self.request_detail.text(),
            self.borrower_input.text(),
            self.code_input.text(),
            self.note_input.text()
        ]
        for col, value in enumerate(values):
            item = QTableWidgetItem(value)
            item.setFont(QFont("Meiryo", 14))
            self.table.setItem(row_position, col, item)

        self.staff_input.clear()
        self.request_detail.clear()
        self.borrower_input.clear()
        self.code_input.clear()
        self.note_input.clear()
        self.request_type.clear_selection()

    def delete_entry(self):
        selected = self.table.selectedItems()
        if selected:
            row = selected[0].row()
            self.table.removeRow(row)

    def init_excel_if_needed(self):
        if not os.path.exists(self.excel_path):
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            ws.append([
                "日付", "担当者名", "申請区分", "申請詳細",
                "借受人氏名", "貸付コード", "特記事項"])
            wb.save(self.excel_path)
        else:
            wb = load_workbook(self.excel_path)
            if self.sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(self.sheet_name)
                ws.append([
                    "日付", "担当者名", "申請区分", "申請詳細",
                    "借受人氏名", "貸付コード", "特記事項"])
                wb.save(self.excel_path)

    def load_existing_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "ファイルを選択", "", "Excel Files (*.xlsx)")
        if path:
            self.excel_path = path
            self.init_excel_if_needed()
            self.load_table()

    def save_data(self):
        if not self.excel_path:
            path, _ = QFileDialog.getSaveFileName(self, "保存ファイルを選択", "申請データ.xlsx", "Excel Files (*.xlsx)")
            if not path:
                QMessageBox.warning(self, "保存中止", "保存先が選択されていません。")
                return
            self.excel_path = path

        self.init_excel_if_needed()
        wb = load_workbook(self.excel_path)
        ws = wb[self.sheet_name]
        ws.delete_rows(2, ws.max_row)

        for row in range(self.table.rowCount()):
            row_data = []
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                row_data.append(item.text() if item else "")
            ws.append(row_data)

        wb.save(self.excel_path)
        QMessageBox.information(self, "保存完了", "Excelに保存されました。")

    def load_table(self):
        wb = load_workbook(self.excel_path)
        ws = wb[self.sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.table.setRowCount(len(rows))
        for i, row in enumerate(rows):
            for j, val in enumerate(row):
                item = QTableWidgetItem(str(val) if val else "")
                item.setFont(QFont("Meiryo", 14))
                self.table.setItem(i, j, item)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    form = LoanFormApp()
    form.show()
    sys.exit(app.exec_())
